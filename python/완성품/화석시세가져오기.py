from selenium import webdriver
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

url2 = 'https://poe.ninja/challenge/fossils'
driver = webdriver.Chrome()
time.sleep(3)
driver.get(url2)
time.sleep(3)
html2 = driver.page_source
soup2 = BeautifulSoup(html2)
data2 = soup2.find('tbody').findAll('tr')
global i
wb = Workbook()
ws1 = wb.active 

for tr in data2[3:]:    
    target1 = tr.findAll('span')[1].text  
    t2src = tr.findAll('span')[4].text  
    target2 = t2src[:len(t2src)-1]
    if target1 == '부식된 화석':
        ws1.cell(row=2, column=1, value=target1) 
        ws1.cell(row=2, column=2, value=target2) 
    if target1 == '인챈트된 화석':  
        ws1.cell(row=3, column=1, value=target1) 
        ws1.cell(row=3, column=2, value=target2)          
    if target1 == '축성된 화석':  
        ws1.cell(row=4, column=1, value=target1) 
        ws1.cell(row=4, column=2, value=target2) 
    if target1 == '완벽한 화석':  
        ws1.cell(row=5, column=1, value=target1) 
        ws1.cell(row=5, column=2, value=target2) 
    if target1 == '전율의 화석':  
        ws1.cell(row=6, column=1, value=target1) 
        ws1.cell(row=6, column=2, value=target2) 
    if target1 == '속박의 화석':  
        ws1.cell(row=7, column=1, value=target1) 
        ws1.cell(row=7, column=2, value=target2) 
    if target1 == '분광 화석':  
        ws1.cell(row=8, column=1, value=target1) 
        ws1.cell(row=8, column=2, value=target2) 
    if target1 == '뾰족한 화석':  
        ws1.cell(row=9, column=1, value=target1) 
        ws1.cell(row=9, column=2, value=target2) 
    if target1 == '에테르 화석':  
        ws1.cell(row=10, column=1, value=target1) 
        ws1.cell(row=10, column=2, value=target2) 
    if target1 == '온전한 화석':  
        ws1.cell(row=11, column=1, value=target1) 
        ws1.cell(row=11, column=2, value=target2) 
    if target1 == '조밀한 화석':  
        ws1.cell(row=12, column=1, value=target1) 
        ws1.cell(row=12, column=2, value=target2) 
    if target1 == '톱니 화석':  
        ws1.cell(row=13, column=1, value=target1) 
        ws1.cell(row=13, column=2, value=target2) 
    if target1 == '빛나는 화석':  
        ws1.cell(row=14, column=1, value=target1) 
        ws1.cell(row=14, column=2, value=target2) 
    if target1 == '금속성 화석':  
        ws1.cell(row=15, column=1, value=target1) 
        ws1.cell(row=15, column=2, value=target2) 
    if target1 == '차디찬 화석':  
        ws1.cell(row=16, column=1, value=target1) 
        ws1.cell(row=16, column=2, value=target2) 
    if target1 == '특이한 화석':  
        ws1.cell(row=17, column=1, value=target1) 
        ws1.cell(row=17, column=2, value=target2)  
wb.save('currency2.xlsx')
driver.close()






